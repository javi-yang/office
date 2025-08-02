import openpyxl
from openpyxl.styles import PatternFill

def extract_margin_data(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    results = []
    green_fill = PatternFill(fill_type="solid", fgColor="00FF00")
    red_fill = PatternFill(fill_type="solid", fgColor="FF0000")

    # 遍历所有sheet
    for ws in wb.worksheets:
        # 查找所有包含"margin"的单元格
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "margin" in str(cell.value):
                    margin_col = cell.column
                    margin_row = cell.row

                    # 向左查找第一个包含"frequency"的单元格
                    freq_col = None
                    for col in range(margin_col - 1, 0, -1):
                        val = ws.cell(row=margin_row, column=col).value
                        if val and "frequency" in str(val):
                            freq_col = col
                            break

                    # 从"margin"单元格向下依次读取
                    scan_row = margin_row + 1
                    while True:
                        margin_val = ws.cell(row=scan_row, column=margin_col).value
                        if margin_val is None:
                            break
                        try:
                            margin_num = float(margin_val)
                        except (TypeError, ValueError):
                            scan_row += 1
                            continue

                        if margin_num < 0 and freq_col:
                            freq_cell = ws.cell(row=scan_row, column=freq_col)
                            freq_val = freq_cell.value
                            freq_cell.fill = green_fill
                            margin_cell = ws.cell(row=scan_row, column=margin_col)
                            margin_cell.fill = red_fill
                            results.append([ws.title, margin_num, freq_val])
                        scan_row += 1

    # 写入output.xlsx
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.append(["Sheet", "Margin", "Frequency"])
    for row in results:
        out_ws.append(row)
    out_wb.save(output_path)
    wb.save(input_path)

if __name__ == "__main__":
    extract_margin_data(r"D:\ALAP\test.xlsx", r"D:\ALAP\output.xlsx")